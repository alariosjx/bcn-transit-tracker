# export/to_excel.py
# Builds data/public/bcn_transit_master.xlsx — reporter-facing deliverable
# Tabs: Summary | Master | Monthly Pivot | YoY | ReadMe | BART | Muni | SMART
# Also writes Datawrapper-ready CSVs to data/public/datawrapper/
#
# Bay City News | Andres Jimenez Larios

import sys
import glob
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, Series
from datetime import datetime, timezone

from merge.config import PROCESSED_DIR, PUBLIC_DIR
from scrapers._utils import log
from export.fiscal_context import FISCAL_CONTEXT, GLOSSARY

PUBLIC_DIR.mkdir(parents=True, exist_ok=True)
DW_DIR = PUBLIC_DIR / "datawrapper"
DW_DIR.mkdir(parents=True, exist_ok=True)

EXCEL_OUT = PUBLIC_DIR / "bcn_transit_master.xlsx"
RAW_DIR   = Path(__file__).resolve().parent.parent / "data" / "raw"

BCN_RED   = "C0392B"
BCN_GRAY  = "F5F5F5"
FLAG_HIGH = "D4EDDA"
FLAG_LOW  = "F8D7DA"
FLAG_WARN = "FFF3CD"


# ── Style helpers ─────────────────────────────────────────────────────────────
def header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        cell.fill      = PatternFill("solid", fgColor=BCN_RED)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def freeze_autofit(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(min(max_len + 4, 80), 12)


def write_df(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=ci, value=col)
    header_row(ws, start_row, len(df.columns))
    for ri, row in enumerate(df.itertuples(index=False), start_row + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci)
            cell.value     = None if (isinstance(val, float) and pd.isna(val)) else val
            cell.font      = Font(name="Arial", size=12)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=BCN_GRAY)


# ── Load SFMTA baselines ──────────────────────────────────────────────────────
def load_muni_baselines() -> dict:
    muni_files = sorted(glob.glob(str(RAW_DIR / "muni_*.csv")), reverse=True)
    if not muni_files:
        return {}
    muni_raw = pd.read_csv(muni_files[0], parse_dates=["date"])
    if "baseline_sfmta" not in muni_raw.columns:
        return {}
    baselines = {}
    for _, row in muni_raw.iterrows():
        month = row["date"].month
        if month not in baselines:
            baselines[month] = int(row["baseline_sfmta"])
    return baselines


# ── Load NTD Financials ───────────────────────────────────────────────────────
def load_financials() -> pd.DataFrame | None:
    fin_files = sorted(glob.glob(str(RAW_DIR / "ntd_financials_*.csv")), reverse=True)
    if not fin_files:
        log.warning("No ntd_financials file found — run scrapers/agencies/ntd_financials.py")
        return None
    return pd.read_csv(fin_files[0])


# ── Load master ───────────────────────────────────────────────────────────────
def load_master() -> pd.DataFrame:
    path = PROCESSED_DIR / "master_monthly.csv"
    if not path.exists():
        raise FileNotFoundError("master_monthly.csv not found. Run merge/build_master.py first.")
    df = pd.read_csv(path, parse_dates=["date"])
    log.info(f"Loaded master: {len(df)} rows")
    return df


# ── Shared stats computation ──────────────────────────────────────────────────
def _agency_stats(adf: pd.DataFrame, agency_id: str, muni_baselines: dict) -> dict:
    """
    Compute latest-month YoY, MoM, and recovery stats for one agency.
    adf must already be non-provisional data, sorted by date.
    Returns an empty dict if adf has no rows.
    """
    if adf.empty:
        return {}
    latest    = adf.iloc[-1]
    prev      = adf.iloc[-2] if len(adf) >= 2 else None
    lat_yr    = latest["date"].year
    lat_mo    = latest["date"].month
    same_ly   = adf[(adf["date"].dt.year == lat_yr - 1) & (adf["date"].dt.month == lat_mo)]
    base_2019 = adf[(adf["date"].dt.year == 2019) & (adf["date"].dt.month == lat_mo)]

    yoy_pct, yoy_text = None, "No prior year data"
    if not same_ly.empty and same_ly["value"].iloc[0] > 0:
        prior    = same_ly["value"].iloc[0]
        yoy_pct  = (latest["value"] - prior) / prior
        arrow    = "▲" if yoy_pct >= 0 else "▼"
        yoy_text = f"{arrow} {abs(yoy_pct):.1%} vs {latest['date'].strftime('%b')} {lat_yr-1} ({int(prior):,} → {int(latest['value']):,})"

    mom_pct, mom_text = None, "No prior month"
    if prev is not None and prev["value"] > 0:
        mom_pct  = (latest["value"] - prev["value"]) / prev["value"]
        arrow    = "▲" if mom_pct >= 0 else "▼"
        mom_text = f"{arrow} {abs(mom_pct):.1%} vs {prev['date'].strftime('%b %Y')} ({int(prev['value']):,} → {int(latest['value']):,})"

    rec_pct, rec_text, rec_val, rec_note = None, "No 2019 baseline", "No 2019 baseline", ""
    if agency_id == "smart":
        rec_text = "N/A — service began 2017, pre-pandemic baseline not comparable"
        rec_val  = "N/A — service began Aug 2017"
    elif agency_id == "muni" and lat_mo in muni_baselines:
        base     = muni_baselines[lat_mo]
        rec_pct  = latest["value"] / base
        rec_text = f"{rec_pct:.1%} of {latest['date'].strftime('%b')} 2019 ({int(base):,} SFMTA baseline)"
        rec_val  = f"{rec_pct:.1%}"
        rec_note = f"SFMTA baseline: {int(base):,}"
    elif not base_2019.empty and base_2019["value"].iloc[0] > 0:
        base     = base_2019["value"].iloc[0]
        rec_pct  = latest["value"] / base
        rec_text = f"{rec_pct:.1%} of {latest['date'].strftime('%b')} 2019 ({int(base):,} baseline)"
        rec_val  = f"{rec_pct:.1%}"
        rec_note = f"2019 baseline: {int(base):,}"

    return {
        "latest"  : latest,
        "prev"    : prev,
        "yoy_pct" : yoy_pct,
        "yoy_text": yoy_text,
        "mom_pct" : mom_pct,
        "mom_text": mom_text,
        "rec_pct" : rec_pct,
        "rec_text": rec_text,
        "rec_val" : rec_val,
        "rec_note": rec_note,
    }


# ── Agency tab helpers ────────────────────────────────────────────────────────
def section_header(ws, row: int, ncols: int, title: str) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor="2C3E50")
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22
    for col in range(2, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill("solid", fgColor="2C3E50")


def kv_row(ws, row: int, label: str, value: str, note: str = "", bold_val: bool = False) -> None:
    lc = ws.cell(row=row, column=1, value=label)
    lc.font      = Font(name="Arial", size=12, bold=True)
    lc.alignment = Alignment(vertical="center")
    vc = ws.cell(row=row, column=2, value=value)
    vc.font      = Font(name="Arial", size=12, bold=bold_val)
    vc.alignment = Alignment(vertical="center")
    if note:
        nc = ws.cell(row=row, column=3, value=note)
        nc.font      = Font(name="Arial", size=11, color="666666", italic=True)
        nc.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 18


def blank_row(ws, row: int) -> None:
    ws.row_dimensions[row].height = 8


# ── Tab 1: Summary ────────────────────────────────────────────────────────────
def tab_summary(wb, master: pd.DataFrame, muni_baselines: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "2C3E50"

    rows = []

    for agency_id in master["agency_id"].unique():
        adf = master[master["agency_id"] == agency_id].sort_values("date")
        if len(adf) < 2:
            continue
        agency_name  = adf["agency_name"].iloc[0]
        adf_complete = adf[~adf["is_provisional"]].copy()
        if adf_complete.empty:
            continue

        s = _agency_stats(adf_complete, agency_id, muni_baselines)
        latest  = s["latest"]
        yoy_pct = s["yoy_pct"]
        mom_pct = s["mom_pct"]

        flags = []
        if yoy_pct is not None:
            if yoy_pct >= 0.10:
                flags.append(f"Strong growth year-over-year (+{yoy_pct:.1%})")
            elif yoy_pct <= -0.10:
                flags.append(f"Notable decline year-over-year ({yoy_pct:.1%})")
        if mom_pct is not None and abs(mom_pct) >= 0.05:
            flags.append(f"Large month-over-month swing ({mom_pct:+.1%})")
        if str(latest.get("flag", "")) in ("REVIEW", "INVESTIGATE"):
            flags.append(f"⚠ Data flag: {latest['flag']} — check agency vs NTD figures")

        rows.append({
            "Agency"           : agency_name,
            "Latest Month"     : latest["date"].strftime("%B %Y"),
            "Ridership"        : f"{int(latest['value']):,}",
            "Source"           : latest["source"],
            "YoY Change"       : s["yoy_text"],
            "Month-over-Month" : s["mom_text"],
            "Recovery vs 2019" : s["rec_text"],
            "Noteworthy"       : " | ".join(flags) if flags else "Within normal range",
        })

    summary_df = pd.DataFrame(rows)
    write_df(ws, summary_df)

    for ri in range(2, len(summary_df) + 2):
        note = ws.cell(row=ri, column=len(summary_df.columns)).value or ""
        if "Strong growth" in note:
            color = FLAG_HIGH
        elif "Notable decline" in note or "INVESTIGATE" in note:
            color = FLAG_LOW
        elif "Provisional" in note or "REVIEW" in note or "Large month" in note:
            color = FLAG_WARN
        else:
            continue
        for ci in range(1, len(summary_df.columns) + 1):
            ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=color)

    freeze_autofit(ws, "A2")
    log.info("Built tab: Summary")


# ── Tab 2: Master ─────────────────────────────────────────────────────────────
def tab_master(wb, master: pd.DataFrame) -> None:
    ws = wb.create_sheet("Master")
    ws.sheet_properties.tabColor = BCN_RED
    df = master.copy()
    df["date"]         = df["date"].dt.strftime("%Y-%m-%d")
    df["variance_pct"] = df["variance_pct"].apply(lambda x: f"{x:.1%}" if pd.notna(x) else "")
    write_df(ws, df)
    freeze_autofit(ws, "A2")
    log.info("Built tab: Master")


# ── Tab 3: Monthly Pivot ──────────────────────────────────────────────────────
def tab_monthly_pivot(wb, master: pd.DataFrame) -> None:
    ws = wb.create_sheet("Monthly Pivot")
    df = master[~master["is_provisional"]].copy()
    df["month_label"] = df["date"].dt.strftime("%b %Y")
    pivot = df.pivot_table(index="agency_name", columns="month_label", values="value", aggfunc="sum")
    pivot = pivot.reindex(
        sorted(pivot.columns, key=lambda x: pd.to_datetime(x, format="%b %Y"), reverse=True), axis=1
    ).reset_index()
    write_df(ws, pivot)
    freeze_autofit(ws, "B2")
    log.info("Built tab: Monthly Pivot")


# ── Tab 4: YoY ────────────────────────────────────────────────────────────────
def tab_yoy(wb, master: pd.DataFrame) -> None:
    ws = wb.create_sheet("Year-over-Year")
    df = master[~master["is_provisional"]].copy()
    df["year"]  = df["date"].dt.year
    df["month"] = df["date"].dt.month
    yoy = df.merge(
        df[["agency_id", "year", "month", "value"]].rename(
            columns={"value": "value_prior", "year": "prior_year"}
        ).assign(year=lambda x: x["prior_year"] + 1),
        on=["agency_id", "year", "month"], how="left"
    )
    yoy["yoy_change"]  = yoy["value"] - yoy["value_prior"]
    yoy["yoy_pct"]     = (yoy["yoy_change"] / yoy["value_prior"]).round(4)
    yoy["yoy_pct_fmt"] = yoy["yoy_pct"].apply(lambda x: f"{x:+.1%}" if pd.notna(x) else "")
    out = yoy[["agency_name", "date", "value", "value_prior", "yoy_change", "yoy_pct_fmt", "source", "flag"]].copy()
    out["date"] = out["date"].dt.strftime("%Y-%m-%d")
    write_df(ws, out.sort_values(["agency_name", "date"], ascending=[True, False]))
    for ri, row in enumerate(yoy.itertuples(index=False), 2):
        if pd.notna(row.yoy_pct):
            if row.yoy_pct >= 0.05:
                color = FLAG_HIGH
            elif row.yoy_pct <= -0.05:
                color = FLAG_LOW
            else:
                continue
            for ci in range(1, len(out.columns) + 1):
                ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=color)
    freeze_autofit(ws, "A2")
    log.info("Built tab: Year-over-Year")


# ── Tab 5: ReadMe ─────────────────────────────────────────────────────────────
def tab_readme(wb) -> None:
    ws = wb.create_sheet("ReadMe")
    ts = datetime.now(timezone.utc).strftime("%B %d, %Y at %H:%M UTC")
    lines = [
        ["BCN Transit Tracker — Bay City News"],
        [f"Last updated: {ts}"],
        [""],
        ["START HERE: Open the Summary tab. It tells you what's notable in plain language."],
        [""],
        ["TABS"],
        ["Summary",              "Plain-language flags. Green = strong growth. Red = decline. Start here."],
        ["Master",               "Full dataset — every month, every agency. All sources and QA flags."],
        ["Monthly Pivot",        "Wide format — one row per agency, columns are months. Good for quick scans."],
        ["Year-over-Year",       "% change vs same month last year. Green = up >5%, Red = down >5%."],
        ["BART / Muni / SMART",  "Per-agency deep dive — at a glance, monthly table, annual financials, charts."],
        ["ReadMe",               "This tab."],
        [""],
        ["DATA FLAGS"],
        ["(blank)",         "Variance between agency and NTD is under 5% — normal."],
        ["REVIEW",          "5–10% variance. Worth a look before publishing."],
        ["INVESTIGATE",     "Over 10% variance. Do not publish without understanding the difference."],
        ["is_provisional",  "Partial month estimate. Numbers will change when month completes."],
        [""],
        ["SOURCES"],
        ["bart.gov",             "BART monthly OD archives (2018–2024) and daily scraper (2025+). Exits metric."],
        ["sfmta.com",            "Muni total boardings. Scraped automatically each weekday at 6 AM PT."],
        ["sonomamarintrain.org", "SMART monthly boardings. Updated manually when SMART publishes new file."],
        ["NTD Monthly",          "National Transit Database. ~2 month lag. Adjusted and finalized."],
        ["NTD Annual",           "Financial data: fares, operating expenses, UPT. Updated annually (~Oct)."],
        [""],
        ["NOTES"],
        ["Muni recovery %",  "Uses SFMTA's own 2019 baseline — methodology matches SFMTA reporting."],
        ["BART recovery %",  "Uses bart.gov 2019 same-month as baseline."],
        ["SMART recovery %", "Not shown — SMART began service Aug 2017, pre-pandemic baseline not comparable."],
        ["Summary tab",      "Shows most recent COMPLETE month only — provisional/partial months excluded."],
        [""],
        ["CONTACT"],
        ["Andres Jimenez Larios | Bay City News"],
        ["GitHub: https://github.com/alariosjx/bcn-transit-tracker"],
    ]
    for ri, line in enumerate(lines, 1):
        for ci, val in enumerate(line, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            if ri == 1:
                cell.font = Font(name="Arial", size=14, bold=True, color=BCN_RED)
            elif ri == 2:
                cell.font = Font(name="Arial", size=10, color="888888")
            elif ri == 4:
                cell.font = Font(name="Arial", size=10, bold=True)
            elif ci == 1 and val and str(val).isupper():
                cell.font = Font(name="Arial", size=10, bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    log.info("Built tab: ReadMe")


# ── Tabs 6-8: Per-agency ──────────────────────────────────────────────────────
def tab_agency(wb, agency_id: str, agency_name: str, master: pd.DataFrame, financials: pd.DataFrame | None, muni_baselines: dict) -> None:
    TAB_NAMES = {"bart": "BART", "muni": "Muni", "smart": "SMART"}
    ws = wb.create_sheet(TAB_NAMES.get(agency_id, agency_name))
    ws.sheet_properties.tabColor = BCN_RED

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 24

    adf         = master[(master["agency_id"] == agency_id) & (~master["is_provisional"])].copy()
    adf         = adf.sort_values("date")
    current_row = 1

    # Title
    title_cell = ws.cell(row=current_row, column=1, value=agency_name)
    title_cell.font      = Font(name="Arial", size=16, bold=True, color=BCN_RED)
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1
    blank_row(ws, current_row); current_row += 1

    # ── Section 1: At a Glance ────────────────────────────────────────────────
    section_header(ws, current_row, 3, "AT A GLANCE"); current_row += 1

    if not adf.empty:
        s      = _agency_stats(adf, agency_id, muni_baselines)
        latest = s["latest"]
        prev   = s["prev"]

        kv_row(ws, current_row, "Latest Complete Month", latest["date"].strftime("%B %Y")); current_row += 1
        kv_row(ws, current_row, "Ridership", f"{int(latest['value']):,}", bold_val=True); current_row += 1

        if s["yoy_pct"] is not None:
            prior = adf[(adf["date"].dt.year == latest["date"].year - 1) & (adf["date"].dt.month == latest["date"].month)]
            prior_val = int(prior["value"].iloc[0]) if not prior.empty else 0
            arrow = "▲" if s["yoy_pct"] >= 0 else "▼"
            kv_row(ws, current_row, "YoY Change",
                   f"{arrow} {abs(s['yoy_pct']):.1%} vs {latest['date'].strftime('%b')} {latest['date'].year-1}",
                   f"({prior_val:,} → {int(latest['value']):,})")
        else:
            kv_row(ws, current_row, "YoY Change", "No prior year data")
        current_row += 1

        if prev is not None and s["mom_pct"] is not None:
            arrow = "▲" if s["mom_pct"] >= 0 else "▼"
            kv_row(ws, current_row, "Month-over-Month",
                   f"{arrow} {abs(s['mom_pct']):.1%} vs {prev['date'].strftime('%b %Y')}",
                   f"({int(prev['value']):,} → {int(latest['value']):,})")
        current_row += 1

        kv_row(ws, current_row, "Recovery vs 2019", s["rec_val"], s["rec_note"])
        current_row += 1

        avg_12 = int(adf.tail(12)["value"].mean())
        kv_row(ws, current_row, "12-Month Rolling Avg", f"{avg_12:,}",
               "Average of last 12 complete months")
        current_row += 1

        blank_row(ws, current_row); current_row += 1

        all_high  = adf.loc[adf["value"].idxmax()]
        all_low   = adf.loc[adf["value"].idxmin()]
        post_adf  = adf[adf["date"] >= "2020-04-01"]
        post_high = post_adf.loc[post_adf["value"].idxmax()] if not post_adf.empty else None
        post_low  = post_adf.loc[post_adf["value"].idxmin()] if not post_adf.empty else None

        kv_row(ws, current_row, "Highest Month (All-Time)",
               f"{int(all_high['value']):,}", all_high["date"].strftime("%B %Y"))
        current_row += 1
        if post_high is not None:
            kv_row(ws, current_row, "Highest Month (Apr 2020+)",
                   f"{int(post_high['value']):,}", post_high["date"].strftime("%B %Y"))
        current_row += 1
        kv_row(ws, current_row, "Lowest Month (All-Time)",
               f"{int(all_low['value']):,}", all_low["date"].strftime("%B %Y"))
        current_row += 1
        if post_low is not None:
            kv_row(ws, current_row, "Lowest Month (Apr 2020+)",
                   f"{int(post_low['value']):,}", post_low["date"].strftime("%B %Y"))
        current_row += 1

    blank_row(ws, current_row); current_row += 1

    # Financials at a glance
    if financials is not None:
        fin_agency = financials[financials["agency_id"] == agency_id].sort_values("report_year")
        if not fin_agency.empty:
            latest_fin = fin_agency.iloc[-1]
            yr         = int(latest_fin["report_year"])
            pre_2019   = fin_agency[fin_agency["report_year"] == 2019]

            kv_row(ws, current_row, f"Fare Revenue ({yr})",
                   f"${int(latest_fin['Fare Revenue']):,}",
                   "Total fares collected from riders (NTD annual)")
            current_row += 1
            kv_row(ws, current_row, f"Operating Expenses ({yr})",
                   f"${int(latest_fin['Operating Expenses']):,}",
                   "Total cost to run the system (NTD annual)")
            current_row += 1
            pre_frr = f"Pre-pandemic (2019): {pre_2019['Fare Recovery Ratio'].iloc[0]:.1%}" if not pre_2019.empty else "Fares ÷ operating expenses"
            kv_row(ws, current_row, f"Fare Recovery Ratio ({yr})",
                   f"{latest_fin['Fare Recovery Ratio']:.1%}", pre_frr)
            current_row += 1
            pre_cpt = f"Pre-pandemic (2019): ${pre_2019['Cost Per Trip'].iloc[0]:.2f}" if not pre_2019.empty else "Operating expenses ÷ annual boardings"
            kv_row(ws, current_row, f"Cost Per Trip ({yr})",
                   f"${latest_fin['Cost Per Trip']:.2f}", pre_cpt)
            current_row += 1

    blank_row(ws, current_row); current_row += 1

    context = FISCAL_CONTEXT.get(agency_id, "")
    if context:
        kv_row(ws, current_row, "⚠ Fiscal Context", context)
        ws.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[current_row].height = 60
        current_row += 1

    blank_row(ws, current_row); current_row += 1

    # ── Section 2: Monthly Ridership (last 24 months) ─────────────────────────
    section_header(ws, current_row, 4, "MONTHLY RIDERSHIP — LAST 24 MONTHS"); current_row += 1

    recent = adf.tail(24).sort_values("date", ascending=False)
    for ci, h in enumerate(["Month", "Ridership", "Source", "YoY Change"], 1):
        c = ws.cell(row=current_row, column=ci, value=h)
        c.font      = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=BCN_RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 1

    for ri, (_, row) in enumerate(recent.iterrows()):
        row_date = row["date"]
        val      = int(row["value"])
        prior_yr = adf[(adf["date"].dt.year == row_date.year - 1) & (adf["date"].dt.month == row_date.month)]
        if not prior_yr.empty:
            p       = prior_yr["value"].iloc[0]
            pct     = (val - p) / p
            arrow   = "▲" if pct >= 0 else "▼"
            yoy_str = f"{arrow} {abs(pct):.1%}"
        else:
            yoy_str = "—"
        fill = PatternFill("solid", fgColor=BCN_GRAY) if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, v in enumerate([row_date.strftime("%B %Y"), f"{val:,}", str(row["source"]), yoy_str], 1):
            c = ws.cell(row=current_row, column=ci, value=v)
            c.font      = Font(name="Arial", size=12)
            c.fill      = fill
            c.alignment = Alignment(vertical="center")
        current_row += 1

    blank_row(ws, current_row); current_row += 1

    # ── Section 3: Annual Financials + Charts ─────────────────────────────────
    if financials is not None:
        fin_agency = financials[financials["agency_id"] == agency_id].sort_values("report_year", ascending=False)
        if not fin_agency.empty:
            section_header(ws, current_row, 6, "ANNUAL FINANCIALS (NTD)"); current_row += 1

            for ci, h in enumerate(["Year", "Fare Revenue", "Operating Expenses", "Fare Recovery", "Cost Per Trip", "Annual Ridership (UPT)"], 1):
                c = ws.cell(row=current_row, column=ci, value=h)
                c.font      = Font(name="Arial", size=12, bold=True, color="FFFFFF")
                c.fill      = PatternFill("solid", fgColor=BCN_RED)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[current_row].height = 30

            fin_data_start = current_row + 1
            current_row   += 1

            for ri, (_, row) in enumerate(fin_agency.iterrows()):
                fill = PatternFill("solid", fgColor=BCN_GRAY) if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                for ci, v in enumerate([
                    int(row["report_year"]),
                    f"${int(row['Fare Revenue']):,}",
                    f"${int(row['Operating Expenses']):,}",
                    f"{row['Fare Recovery Ratio']:.1%}",
                    f"${row['Cost Per Trip']:.2f}",
                    f"{int(row['Annual Ridership (UPT)']):,}",
                ], 1):
                    c = ws.cell(row=current_row, column=ci, value=v)
                    c.font      = Font(name="Arial", size=12)
                    c.fill      = fill
                    c.alignment = Alignment(vertical="center")
                current_row += 1

            fin_data_end = current_row - 1
            blank_row(ws, current_row); current_row += 1

            # Write raw numeric backing data in hidden cols G-K (ascending by year for charts)
            fin_asc   = fin_agency.sort_values("report_year")
            chart_row = fin_data_start
            for _, row in fin_asc.iterrows():
                ws.cell(row=chart_row, column=26, value=int(row["report_year"]))
                ws.cell(row=chart_row, column=27, value=int(row["Fare Revenue"]))
                ws.cell(row=chart_row, column=28, value=int(row["Operating Expenses"]))
                ws.cell(row=chart_row, column=29, value=float(row["Cost Per Trip"]))
                ws.cell(row=chart_row, column=30, value=int(row["Annual Ridership (UPT)"]))
                chart_row += 1


            years_ref     = Reference(ws, min_col=26, min_row=fin_data_start, max_row=fin_data_end)
            fare_ref      = Reference(ws, min_col=27, min_row=fin_data_start, max_row=fin_data_end)
            opex_ref      = Reference(ws, min_col=28, min_row=fin_data_start, max_row=fin_data_end)
            cpt_ref       = Reference(ws, min_col=29, min_row=fin_data_start, max_row=fin_data_end)
            ridership_ref = Reference(ws, min_col=30, min_row=fin_data_start, max_row=fin_data_end)

            # Chart 1: Fare Revenue vs Operating Expenses
            c1 = LineChart()
            c1.title        = "Fare Revenue vs Operating Expenses"
            c1.style        = 10
            c1.height       = 10
            c1.width        = 20
            c1.y_axis.title = "$"
            s1 = Series(fare_ref, title="Fare Revenue"); s1.smooth = True
            s2 = Series(opex_ref, title="Operating Expenses"); s2.smooth = True
            c1.append(s1); c1.append(s2)
            c1.set_categories(years_ref)
            ws.add_chart(c1, f"J{fin_data_start}")

            # Chart 2: Annual Ridership
            c2 = LineChart()
            c2.title        = "Annual Ridership (UPT)"
            c2.style        = 10
            c2.height       = 10
            c2.width        = 20
            c2.y_axis.title = "Boardings"
            s3 = Series(ridership_ref, title="Annual Ridership"); s3.smooth = True
            c2.append(s3)
            c2.set_categories(years_ref)
            ws.add_chart(c2, f"J{fin_data_start + 16}")

            # Chart 3: Cost Per Trip
            c3 = LineChart()
            c3.title        = "Cost Per Trip ($)"
            c3.style        = 10
            c3.height       = 10
            c3.width        = 20
            c3.y_axis.title = "$ per trip"
            s4 = Series(cpt_ref, title="Cost Per Trip"); s4.smooth = True
            c3.append(s4)
            c3.set_categories(years_ref)
            ws.add_chart(c3, f"J{fin_data_start + 32}")

    # ── Section 4: Glossary ───────────────────────────────────────────────────
    section_header(ws, current_row, 3, "METRIC DEFINITIONS"); current_row += 1
    for term, definition in GLOSSARY:
        kv_row(ws, current_row, term, definition)
        current_row += 1

    ws.freeze_panes = "A2"
    log.info(f"Built tab: {TAB_NAMES.get(agency_id, agency_name)}")


# ── Datawrapper CSVs ──────────────────────────────────────────────────────────
def write_datawrapper_csvs(master: pd.DataFrame, muni_baselines: dict) -> None:
    base = master[~master["is_provisional"]].copy()
    agencies = {
        "bart" : "Bay Area Rapid Transit",
        "muni" : "San Francisco Municipal Railway",
        "smart": "Sonoma-Marin Area Rail Transit",
        "actransit": "AC Transit",
        "caltrain" : "Caltrain",
    }

    for agency_id, agency_name in agencies.items():
        adf = base[base["agency_name"] == agency_name].copy()
        if adf.empty:
            continue

        # 1. Timeseries
        ts = adf[["date", "agency_name", "value", "source"]].copy()
        if agency_id == "muni" and muni_baselines:
            ts = ts[ts["source"] != "NTD"].copy()
            baseline_rows = pd.DataFrame([
                {"date": pd.Timestamp(f"2019-{m:02d}-01"), "agency_name": agency_name,
                 "value": v, "source": "sfmta.com"}
                for m, v in sorted(muni_baselines.items())
            ])
            ts = pd.concat([baseline_rows, ts], ignore_index=True)
        ts["date"] = ts["date"].dt.strftime("%Y-%m-%d")
        ts.sort_values("date").to_csv(DW_DIR / f"{agency_id}_monthly_timeseries.csv", index=False)

        # 2. YoY
        yoy = adf.copy()
        yoy["year"]        = yoy["date"].dt.year
        yoy["month"]       = yoy["date"].dt.month
        yoy["month_label"] = yoy["date"].dt.strftime("%b")
        current_year = yoy["year"].max()
        prior_year   = current_year - 1
        yoy_wide = yoy[yoy["year"].isin([prior_year, current_year])].pivot_table(
            index=["month", "month_label"], columns="year", values="value", aggfunc="sum"
        ).reset_index()
        yoy_wide.columns = [str(c) for c in yoy_wide.columns]
        for yr in [str(prior_year), str(current_year)]:
            if yr in yoy_wide.columns:
                yoy_wide[yr] = pd.to_numeric(yoy_wide[yr], errors="coerce").apply(
                    lambda x: int(x) if pd.notna(x) else ""
                )
        month_order = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        yoy_wide["month_label"] = pd.Categorical(yoy_wide["month_label"], categories=month_order, ordered=True)
        cols = ["month_label", str(prior_year), str(current_year)]
        cols = [c for c in cols if c in yoy_wide.columns]
        yoy_wide.sort_values("month_label")[cols].to_csv(DW_DIR / f"{agency_id}_yoy_comparison.csv", index=False)

        # 3. Recovery tracker (skip SMART)
        if agency_id == "smart":
            continue
        rec = adf.copy()
        rec["year"]  = rec["date"].dt.year
        rec["month"] = rec["date"].dt.month
        if agency_id == "muni" and muni_baselines:
            muni_bl_df = pd.DataFrame([{"month": m, "baseline_2019": v} for m, v in muni_baselines.items()])
            rec = rec[rec["year"] >= 2020].merge(muni_bl_df, on="month", how="left")
        else:
            ntd_baseline = rec[rec["year"] == 2019][["month", "value"]].rename(columns={"value": "baseline_2019"})
            rec = rec[rec["year"] >= 2020].merge(ntd_baseline, on="month", how="left")
        rec["recovery_pct"] = (rec["value"] / rec["baseline_2019"] * 100).round(1)
        rec["date"]         = rec["date"].dt.strftime("%Y-%m-%d")
        rec[["agency_name", "date", "value", "baseline_2019", "recovery_pct"]].to_csv(
            DW_DIR / f"{agency_id}_recovery_tracker.csv", index=False
        )
        log.info(f"Wrote Datawrapper CSVs for {agency_id}")

    log.info(f"Wrote all Datawrapper CSVs → {DW_DIR}")


# ── Main ──────────────────────────────────────────────────────────────────────
def run():
    log.info("── to_excel.py starting ──")
    master         = load_master()
    financials     = load_financials()
    muni_baselines = load_muni_baselines()

    wb = Workbook()
    wb.remove(wb.active)

    tab_summary(wb, master, muni_baselines)
    tab_master(wb, master)
    tab_monthly_pivot(wb, master)
    tab_yoy(wb, master)
    tab_readme(wb)

    tab_agency(wb, "bart",             "Bay Area Rapid Transit",                    master, financials, muni_baselines)
    tab_agency(wb, "muni",             "San Francisco Municipal Railway",           master, financials, muni_baselines)
    tab_agency(wb, "smart",            "Sonoma-Marin Area Rail Transit",            master, financials, muni_baselines)
    tab_agency(wb, "actransit",        "AC Transit",                                master, financials, muni_baselines)
    tab_agency(wb, "caltrain",         "Caltrain",                                  master, financials, muni_baselines)
    tab_agency(wb, "sfbayferry",       "SF Bay Ferry",                              master, financials, muni_baselines)
    tab_agency(wb, "samtrans",         "SamTrans",                                  master, financials, muni_baselines)
    tab_agency(wb, "vta",              "VTA",                                       master, financials, muni_baselines)
    tab_agency(wb, "ggferry",          "Golden Gate Ferry",                         master, financials, muni_baselines)
    tab_agency(wb, "ggbus",            "Golden Gate Bus",                           master, financials, muni_baselines)
    tab_agency(wb, "marin",            "Marin Transit",                             master, financials, muni_baselines)
    tab_agency(wb, "napa",             "Napa Valley Transportation Authority",      master, financials, muni_baselines)
    tab_agency(wb, "vallejo",          "Vallejo Transit",                           master, financials, muni_baselines)
    tab_agency(wb, "countyconnection", "County Connection",                         master, financials, muni_baselines)
    tab_agency(wb, "westcat",          "WestCAT",                                   master, financials, muni_baselines)
    tab_agency(wb, "tridelta",         "Tri Delta Transit",                         master, financials, muni_baselines)
    tab_agency(wb, "wheels",           "Wheels (LAVTA)",                            master, financials, muni_baselines)
    tab_agency(wb, "unioncity",        "Union City Transit",                        master, financials, muni_baselines)
    tab_agency(wb, "alamedaferry",     "Alameda Ferry",                             master, financials, muni_baselines)
    tab_agency(wb, "santarosa",        "Santa Rosa CityBus",                        master, financials, muni_baselines)
    tab_agency(wb, "fairfield",        "Fairfield-Suisun Transit",                  master, financials, muni_baselines)
    tab_agency(wb, "vacaville",        "Vacaville City Coach",                      master, financials, muni_baselines)
    tab_agency(wb, "petaluma",         "Petaluma Transit",                          master, financials, muni_baselines)
    tab_agency(wb, "mst",              "Monterey-Salinas Transit",                  master, financials, muni_baselines)
    tab_agency(wb, "santacruz",        "Santa Cruz Metro",                          master, financials, muni_baselines)
    tab_agency(wb, "sjrtd",            "San Joaquin RTD",                           master, financials, muni_baselines)
    tab_agency(wb, "ace",              "Altamont Corridor Express",                 master, financials, muni_baselines)

    wb.save(EXCEL_OUT)
    log.info(f"Saved Excel: {EXCEL_OUT}")

    write_datawrapper_csvs(master, muni_baselines)

    for agency_id in master["agency_id"].unique():
        adf = master[(master["agency_id"] == agency_id) & (~master["is_provisional"])].copy()
        adf["date"] = adf["date"].dt.strftime("%Y-%m-%d")
        adf.to_csv(PUBLIC_DIR / f"{agency_id}_monthly.csv", index=False)
        log.info(f"Wrote: {agency_id}_monthly.csv")

    log.info("── to_excel.py complete ──")


if __name__ == "__main__":
    run()